import csv
import io
import json
from pathlib import Path

import fitz
from docx import Document as DocxDocument
from openpyxl import load_workbook


def extract_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".txt", ".md"}:
        return path.read_text(encoding="utf-8")
    if suffix == ".json":
        return json.dumps(json.loads(path.read_text(encoding="utf-8")), indent=2)
    if suffix == ".csv":
        return path.read_text(encoding="utf-8")
    if suffix == ".pdf":
        with fitz.open(path) as doc:
            return "\n".join(page.get_text() for page in doc)
    if suffix == ".docx":
        doc = DocxDocument(path)
        return "\n".join(p.text for p in doc.paragraphs)
    if suffix == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        lines = []
        for sheet in workbook.worksheets:
            lines.append(f"Sheet: {sheet.title}")
            lines.extend(",".join("" if v is None else str(v) for v in row) for row in sheet.iter_rows(values_only=True))
        return "\n".join(lines)
    raise ValueError(f"Unsupported file type: {suffix}")


def csv_text(rows: list[dict]) -> str:
    if not rows:
        return ""
    stream = io.StringIO()
    writer = csv.DictWriter(stream, fieldnames=rows[0].keys())
    writer.writeheader()
    writer.writerows(rows)
    return stream.getvalue()
